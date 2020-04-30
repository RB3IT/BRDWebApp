## Builtin
import io
import itertools
## Third Party
import openpyxl
from openpyxl.worksheet import table as xltable
from openpyxl  import styles as xlstyle, utils as xlutils
from openpyxl.formatting import rule as xlformatrule
## This Module
from . import methods
from . import subclasses

## ... The point was to have the strftimes and labels in one unified position...
## Unfortunately, STRINGFORMAT = "{kwarg.method(*args,**kw)}" doesn't work because
## the Formatter expects kwarg to have an attribute called "method(*args,**kw)"
## This doesn't happen with f-strings because they expect to evaluate the contents
## (i.e.- f"{kwarg.method(*args,**kw)}" -> "{result}".format(result = kwarg.method(*args,**kw)) -> "result")
def MONTHTABLENAME(month):
    return f"{month.strftime('%B_%Y')}_Inventory"
def  MONTHTITLE(month):
    return f"{month.strftime('%B %Y')}"

####################################################################
"""-----------------------------------------------------------------
                            STYLE SETUP
-----------------------------------------------------------------"""
####################################################################

TITLESTYLE = xlstyle.NamedStyle(name='TITLESTYLE')
TITLESTYLE.font = xlstyle.Font(bold = True,size = 16)
TITLESTYLE.alignment = xlstyle.Alignment(horizontal = "center")

TOTALCELLSTYLE = xlstyle.NamedStyle(name="TOTALCELLSTYLE")
TOTALCELLSTYLE.font = xlstyle.Font(bold = True, size = 12, color = 'FFFFFF')
ALLTHICKBORDERS = xlstyle.Border(**{side:xlstyle.Side(border_style='thick',color = '000000') for side in ['left','right','top','bottom']})
TOTALCELLSTYLE.border = ALLTHICKBORDERS
TOTALCELLSTYLE.fill = xlstyle.PatternFill(fill_type = "solid", start_color = 'BFBFBF')

MONTHTABLESTYLE = xltable.TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)

SUMMARYTABLESTYLE = xltable.TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=False, showColumnStripes=False)

def TOTALSCONDITIONALSTYLE(worksheet,cellrange,secondcolumn,firstrow):
    lessthanrule = xlformatrule.FormulaRule(formula=[f"{xlutils.get_column_letter(secondcolumn)}{firstrow}<{xlutils.get_column_letter(secondcolumn-1)}{firstrow}"],fill = xlstyle.PatternFill(fill_type="solid",start_color="DEEBF6",end_color="DEEBF6"))
    worksheet.conditional_formatting.add(cellrange,lessthanrule)
    greaterthanrule = xlformatrule.FormulaRule(formula=[f"{xlutils.get_column_letter(secondcolumn)}{firstrow}>{xlutils.get_column_letter(secondcolumn-1)}{firstrow}"],fill = xlstyle.PatternFill(fill_type="solid",start_color="FFF3CB",end_color="FFF3CB"))
    worksheet.conditional_formatting.add(cellrange,greaterthanrule)

####################################################################
"""-----------------------------------------------------------------
                              Methods
-----------------------------------------------------------------"""
####################################################################

def autowidth(worksheet,column):
    """ Sets the width of a column to _approximately_ the correct width for its content
    
    Adapted from http://stackoverflow.com/a/39530676
    """
    max_length = 0
    try:
        columnname = column[0].column_letter # Get the column name
    except:
        ## Merged cell
        columnname = xlutils.get_column_letter(column[0].column)
    for cell in column:
        try: # Necessary to avoid error on empty cells
            max_length = max(len(str(cell.value)),max_length)
        except: pass
    adjusted_width = (max_length + 2) * 1.2
    ## Patch 3/6/19: Excel asking for string here for some reason
    worksheet.column_dimensions[columnname].width = adjusted_width

####################################################################
"""-----------------------------------------------------------------
                              Export
-----------------------------------------------------------------"""
####################################################################

def createexceldownload(start,end):
    workbook = openpyxl.Workbook()
    ## Remove default sheet
    try:
        ## Depricated call saved temporarily for reference
        ## workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
        del workbook["Sheet"]
    except: pass
    for month in methods.itermonths(start,end):
        inventory = methods.getinventorybymonth(month)
        ## Mutating method (sets .cost on existing items)
        methods.getcostforinventory(*inventory)
        ## Mutating (adds Worksheets)
        createexcelmonthsheet(workbook,month,inventory)

    createexcelbyitemsummary(workbook,start,end)

    ## Save
    out = io.BytesIO()
    ## Debugging
    ## workbook.save('testoutput.xlsx')
    workbook.save(out)
    out.seek(0)
    return out


def createexcelmonthsheet(workbook,month,inventory):
    """ Creates an Excel Inventory Report for the given Month with the given Inventory as the first sheet in the workbook.

    Workbook should be an openpxyl Workbook.
    Month is a datetime object.
    inventory should be a prefiltered list of subclasses.InventoryOutput objects to include.
    """
    worksheet = workbook.create_sheet(title=MONTHTITLE(month=month),index=0)

    TABLENAME = MONTHTABLENAME(month = month)

    ## Sheet Title
    worksheet.merge_cells("A1:E1")
    titlecell = worksheet['A1']
    titlecell.value = MONTHTITLE(month = month)
    titlecell.style = TITLESTYLE

    ## Total Line
    # Title
    totalcell = worksheet['B2']
    totalcell.value = "Total Inventory"
    totalcell.style = TOTALCELLSTYLE
    

    # Value
    totalvaluecell = worksheet['C2']
    totalvaluecell.value = f"=SUM({TABLENAME}[Total])"
    totalvaluecell.font = xlstyle.Font(bold=True, size = 12)

    ## Table
    # Header
    for column,header in enumerate(['ItemID','Description','Count','Cost','Total'],start=1):
        worksheet.cell(column=column, row=4, value = header)

    # Inventory
    ## For each item, make a new row
    itemrow = 5
    for row,item in enumerate(inventory,start=itemrow):
        ## for each value, use a new column
        values = [item.itemid, item.description, item.quantity]
        for column,value in enumerate(values,start=1):
            ## set value
            worksheet.cell(row = row, column = column, value = value)
        currencies = [item.cost]
        for column,value in enumerate(currencies,start=len(values)+1):
            ## set value
            cell = worksheet.cell(row = row, column = column, value = value)
            cell.number_format='$#,##0.00'
        ## Add total
        cell = worksheet.cell(row = row, column = len(values)+len(currencies)+1,
                       value = f'={TABLENAME}[[#This Row],[Cost]]*{TABLENAME}[[#This Row],[Count]]')
        cell.number_format='$#,##0.00'

    ## Define Table
    table = xltable.Table(displayName=TABLENAME, ref = f"A4:E{len(inventory)+4}")
    table.tableStyleInfo = MONTHTABLESTYLE
    worksheet.add_table(table)
    autowidth(worksheet,list(worksheet.columns)[0])
    autowidth(worksheet,list(worksheet.columns)[1])
    return worksheet


def createexcelbyitemsummary(workbook,start,end):
    """ Create an Excel Sheet at the first index that lists Total Costs and Item Totals for each month between start and end.
    
    Note that this sheet only contains references to other tables that may or may-not exist; they are prepended
    with IFERROR to maintain visual integrity.
    workbook should be an openpyxl workbook object.
    start and end should be datetime objects.
    Returns a reference to the created Worksheet (which had already been inserted)
    """

    worksheet = workbook.create_sheet(title="Year Costs Summary",index=0)

    months = list(methods.itermonths(start,end))
    items = list(set(
        itertools.chain.from_iterable([methods.getincludeditems(month) for month in months])
        ))
    items = sorted(
        sorted(items,key = lambda item: item.description),
        key = lambda item: item.itemindex
        )
    LASTCOLUMN = xlutils.get_column_letter(2+len(months))
    SUMMARYTABLENAME = "Year_Costs_Summary"
    ITEMTABLENAME = "Year_Items_Costs_Summary"

    ## Sheet Title
    ## Title will be as long as daterange
    worksheet.merge_cells(f"A1:{LASTCOLUMN}1")
    titlecell = worksheet['A1']
    titlecell.value = f"Year Costs Summary"
    titlecell.style = TITLESTYLE


    ## Totals Row
    totalrow = 3
    totalcolumn = 3
    for i,month in enumerate(months):
        ## Header
        worksheet.cell(row = totalrow, column = totalcolumn + i, value = MONTHTITLE(month=month))
        ## Value
        cell = worksheet.cell(row = totalrow+1, column = totalcolumn + i, value = f'=IFERROR(SUM({MONTHTABLENAME(month=month)}[Total]),"")')
        cell.number_format='$#,##0.00'

    # Define Table
    table = xltable.Table(displayName=SUMMARYTABLENAME, ref = f"C3:{LASTCOLUMN}4")
    table.tableStyleInfo = SUMMARYTABLESTYLE
    worksheet.add_table(table)

    ## Items Table
    # Header
    headerrow = 6
    for column,header in enumerate(['ItemID','Description']+[MONTHTITLE(month = month) for month in months],start=1):
        cell = worksheet.cell(column=column, row=headerrow, value = header)
        cell.number_format='$#,##0.00'

    # Items
    firstitemrow = 7
    for i,item in enumerate(items):
        ## Basic info (Note, if this grows any more than this, use iterator)
        worksheet.cell(row = firstitemrow+i, column = 1, value = item.itemid)
        worksheet.cell(row = firstitemrow+i, column = 2, value = item.description)
        ## Months
        for m, month in enumerate(months,start =1):
            ## [#Data] must be used for autogeneration; don't know why, just know
            ## that the worksheet will trigger a ?NameError
            cell = worksheet.cell(row = firstitemrow+i, column = 2+m, value = f'=IFERROR(VLOOKUP({ITEMTABLENAME}[[#This Row],[ItemID]],{MONTHTABLENAME(month=month)}[#Data],5,0),"--")')
            cell.number_format='$#,##0.00'

    # Define Table
    table = xltable.Table(displayName=ITEMTABLENAME, ref = f"A{headerrow}:{LASTCOLUMN}{firstitemrow+len(items)-1}")
    table.tableStyleInfo = SUMMARYTABLESTYLE
    worksheet.add_table(table)

    ## Adjust item columns
    for column in list(worksheet.columns)[:2]:
        autowidth(worksheet,column)

    for column in list(worksheet.columns)[2:]:
        try:
            columnname = column[0].column_letter # Get the column name
        except:
            ## Merged cell
            columnname = xlutils.get_column_letter(column[0].column)
        worksheet.column_dimensions[columnname].width = 16.5

    ## Add Conditional Formatting
    ## Only conditional format when we have 2 or more months to compare
    if len(months)>1:
        cellrange = f"D{firstitemrow}:{LASTCOLUMN}{firstitemrow+len(items)-1}"
        TOTALSCONDITIONALSTYLE(worksheet=worksheet,cellrange=cellrange,secondcolumn=4,firstrow=firstitemrow)
    
    return worksheet